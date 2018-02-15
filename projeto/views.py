from collections import namedtuple

from django.http import HttpResponse
from django.shortcuts import render, redirect
#from django.conf import settings
#from django.core.files.storage import FileSystemStorage
from projeto.forms import FormGerarRelatoriosFinanceiros
from carregaarquivo.forms import FormExibePlanilha
from carregaarquivo.models import Arquivo, FuncaoArquivo, TipoArquivo
from orcamento.models import Orcamento
from pagamento.models import Pagamento
from .models import Projeto, Tarefa
from utils.utilitarios import *
from openpyxl import Workbook
from openpyxl.styles import Alignment
import datetime
from carregaarquivo.views import Carregaarquivo, ImportaPlanilha
import locale


# Create your views here.

def index(request):
    projetos = Projeto.objects.all()
    return render(request, 'projeto/listaprojetos.html', {'projetos':projetos,})

def tarefas(request, id_projeto):
    projeto = Projeto.objects.get(pk=id_projeto)
    tarefas = Tarefa.objects.filter(projeto=projeto).order_by('status','prioridade','data_criacao')
    pendentes = {}
    em_andamento = {}
    concluidas = {}
    if tarefas:
        cor_prioridade = ""
        for tarefa in tarefas:
            envolvidos = "Responsável: " + str(tarefa.responsavel) + "\nEnvolvidos: " + tarefa.envolvidos
            if tarefa.prioridade == 0:
                cor_prioridade = cores('critica')
            elif tarefa.prioridade == 1:
                cor_prioridade = cores('alerta')
            else:
                cor_prioridade = cores('padrao')
            item = {'descricao': tarefa.desricao_tarefa,
                    'envolvidos':envolvidos,
                    'meta':tarefa.meta,
                    'prioridade':tarefa.prioridade,
                    'data_criacao':tarefa.data_criacao,
                    'data_limite':tarefa.data_conclusao,
                    'cor_prioridade':cor_prioridade}
            if tarefa.status == 0:
                pendentes[tarefa.pk] = item
            elif tarefa.status == 1:
                em_andamento[tarefa.pk] = item
            else:
                concluidas[tarefa.pk] = item

    return render(request, 'projeto/exibe_tarefas.html', {'projeto':projeto,
                                                          'pendentes':pendentes,
                                                          'em_andamento':em_andamento,
                                                          'concluidas':concluidas})

def editatarefa(request, id_tarefa):
    pass

def tarefaemandamento(request, id_tarefa):
    pass

def tarefaconcluida(request, id_tarefa):
    pass

def tarefapendente(request, id_tarefa):
    pass

def carrega_orcamento(request, id_projeto):
    '''
    View que constroi o formulario de upload das planilhas.

    :param request:
    :return: objeto upload_engine contendo formulario para upload de planilhas
    '''
    upload_engine = Carregaarquivo()
    upload_engine.objeto_associado = 'Projeto.pk=[{}]'.format(id_projeto)
    if not upload_engine.model_form_upload(request):
        return redirect('planilhascarregadas', id_projeto=id_projeto)
    return render(request, 'projeto/carregaorcamento.html', {'upload':upload_engine})

def planilhascarregadas(request, id_projeto):
    '''
    Recupera a lista das planilhas que foram carregadas para a pasta MEDIA.
    :param request:
    :return: lista dos registros contendo os arquivos de planilha
    '''
    projeto = Projeto.objects.get(pk=id_projeto)
    arquivos_planilha = TipoArquivo.objects.filter(extencao_arquivo__in = ['xls', 'xlsx'])
    funcao_orcamento = FuncaoArquivo.objects.filter(sigla_funcao='PLO')
    arquivos = Arquivo.objects.filter(tipo_arquivo__in = list(arquivos_planilha),
                                      funcao_arquivo = funcao_orcamento[0],
                                      objeto_associado__contains = 'Projeto.pk=[{}]'.format(id_projeto))
    return render(request, 'projeto/lista_planilhas_carregadas.html', {'arquivos':arquivos, 'projeto':projeto})

def exibir_planilha(request, id_projeto, id_arquivo):
    '''
    Monta e exibe uma tabela a partir de uma planilha excel carregada previamente carregada na pasta MEDIA.

    :param request: class HttpRequest
    :param id_arquivo: PK proveniente da tabela de dados dos arquivos carregados na pasta MEDIA
    :return: uma STR contendo a tabela a ser renderizada no template
    '''

    projeto = Projeto.objects.get(pk=id_projeto)
    planilha = ""
    arquivo = Arquivo.objects.get(pk=id_arquivo)

    if arquivo:

        planilha_importada = False
        desconsiderar_linha = ""

        ipp = ImportaPlanilha()
        if arquivo.filtro:
            ipp.ativa_filtro(arquivo.filtro)

        if request.method == 'POST':
            formulario = FormExibePlanilha(request.POST, initial={'acao': 'Vizualizar'})
            if formulario.is_valid():
                ipp.set_linha_final(int(formulario.cleaned_data['linha_final']))
                desconsiderar_linha = formulario.cleaned_data['desconsiderar_linhas']
                ipp.set_desconsidera_linhas(desconsiderar_linha)
                if formulario.cleaned_data['acao'] == 'Importar':
                    print("*****Agora vou importar com a porra toda*****")
        else:
            if arquivo.filtro:
                desconsiderar_linha = arquivo.filtro.excecao_linhas
                formulario = FormExibePlanilha({'linha_final':arquivo.filtro.linha_final,
                                                'desconsiderar_linhas':desconsiderar_linha},
                                               initial={'acao': 'Vizualizar'})
            else:
                formulario = FormExibePlanilha(initial={'acao': 'Vizualizar'})

        planilha_importada = ipp.importa_planilha(arquivo)

        tb = TabelaHTML()
        tb.class_padrao = 'table table-bordered'
        if planilha_importada:
            tb.numera_linhas = True
            tb.inicio_contagem = ipp.get_linha_inicial()
            tb.intevalo_marcacao_cor_especial = lista_de_intervalos(desconsiderar_linha)
            tb.cor_especial = cores('alerta')
            planilha = tb.gerar_tabela(planilha_importada)
        else:
            tb.numera_linhas = False
            planilha = 'Não foi possível importar a planilha. Verifique se o filtro está ajustado ou se o arquivo está presente'

    return render(request, 'projeto/exibe_planilha.html', {'planilha':planilha,'arquivo':arquivo,
                                                           'formulario':formulario, 'projeto':projeto})

def relatorios(request, id_projeto):
    '''
    Produz o relatorio financeiro e a conciliacao bancaria gerando um arquivo excel com as informacoes

    :param request:
    :param id_projeto:
    :return: arquivo excel contendo relatorio financeiro e conciliacao bancaria
    '''
    projeto = Projeto.objects.get(pk=id_projeto)

    if request.method == 'POST':
        formulario = FormGerarRelatoriosFinanceiros(request.POST)
        if formulario.is_valid():
            pagamentos = Pagamento.objects.filter(id_projeto=projeto)
            if pagamentos:
                DetalhesFinaceiro = namedtuple('DetalhesFinaceiro', ('itens', 'descricao', 'tipo_contratacao', 'id_docs_fiscais',
                                                                     'valor', 'nr_doc_forma'))
                DetalhesConciliacao = namedtuple('DetalhesConciliacao', ('nr_doc_forma', 'data_emissao', 'beneficiario', 'valor'))
                relatorio_financeiro = []
                relatorio_financeiro.append(DetalhesFinaceiro('Rubrica','Descrição','Tipo de contratação','Documento fiscal',
                                                              'Valor pago','Cheque (nº)'))
                relatorio_conciliacao = []
                relatorio_conciliacao.append(DetalhesConciliacao('Nº CHEQUE', 'DATA DE EMISSÃO', 'FORNECEDOR', 'VALOR (R$)'))
                locale.setlocale(locale.LC_ALL, '')
                for pagamento in pagamentos:
                    if pagamento.formas_pagamento and pagamento.formas_comprovacao:
                        relatorio_financeiro.append(DetalhesFinaceiro(";".join(pagamento.lista_desc_itens_pagamento),
                                                                      pagamento.descricao_docs_comprovacao,
                                                                      pagamento.id_pessoa.tipo_pessoa,
                                                                      pagamento.identificacao_docs_comprovacao,
                                                                      pagamento.valor_liquido,
                                                                      pagamento.identificacao_formas_pagamento))
                        relatorio_conciliacao.append(DetalhesConciliacao(pagamento.identificacao_formas_pagamento,
                                                                         pagamento.formas_pagamento[0].data_emissao.strftime("%d/%m/%y"),
                                                                         str(pagamento.id_pessoa), pagamento.valor_liquido))
                    else:
                        relatorio_financeiro.append(DetalhesFinaceiro(False, False, False, False, False, False))
                        relatorio_conciliacao.append(DetalhesConciliacao(False, False, False, False))

                wb = Workbook()
                ws_financeiro = wb.create_sheet('Relatório Financeiro', 0)
                ws_conciliacao = wb.create_sheet('Conciliação Bancária', 1)
                AtributosCelula = namedtuple('AtributosCelula', ('largura','al_vertical','al_horizontal','wrap'))
                dim_align_cols = (AtributosCelula(50,'top','left', True), AtributosCelula(50,'top','left',True),
                                  AtributosCelula(15,'center','center', True), AtributosCelula(15,'center','center',True),
                                  AtributosCelula(10,'center','center', True), AtributosCelula(10,'center','center',True))

                for nr_linha, linha in enumerate(relatorio_financeiro):
                    idx_col_A = ord('A')
                    for nr_coluna, coluna in enumerate(linha):
                        celula = ws_financeiro.cell(row=nr_linha+1, column=nr_coluna+1, value=coluna)
                        celula.alignment = Alignment(vertical=dim_align_cols[nr_coluna].al_vertical,
                                                     horizontal=dim_align_cols[nr_coluna].al_horizontal,
                                                     wrapText=dim_align_cols[nr_coluna].wrap)
                        ws_financeiro.column_dimensions[chr(idx_col_A)].width = dim_align_cols[nr_coluna].largura
                        idx_col_A +=1
                l_final = 1
                dim_align_cols = (AtributosCelula(10,'top','center',True), AtributosCelula(10,'top','center',True),
                                  AtributosCelula(50,'top','left',True), AtributosCelula(10,'top','center', True))
                for nr_linha, linha in enumerate(relatorio_conciliacao):
                    idx_col_A = ord('A')
                    for nr_coluna, coluna in enumerate(linha):
                        celula = ws_conciliacao.cell(row=nr_linha+1, column=nr_coluna+1, value=coluna)
                        celula.alignment = Alignment(vertical=dim_align_cols[nr_coluna].al_vertical,
                                                     horizontal=dim_align_cols[nr_coluna].al_horizontal,
                                                     wrapText=dim_align_cols[nr_coluna].wrap)
                        ws_conciliacao.column_dimensions[chr(idx_col_A)].width = dim_align_cols[nr_coluna].largura
                        l_final = nr_linha + 1
                l_final += 1
                ws_conciliacao.cell(row=l_final, column=1, value='TOTAL')
                ws_conciliacao.cell(row=l_final, column=4, value='=SUM(D2:D{})'.format(l_final-1))
                ws_conciliacao.merge_cells(start_row=l_final, start_column=1, end_row=l_final, end_column=3)

                l_final += 1
                ws_conciliacao.cell(row=l_final+1, column=1, value='CRÉDITO CONCEDIDO (R$) [A]')
                ws_conciliacao.merge_cells(start_row=l_final+1, start_column=1, end_row=l_final+1, end_column=3)
                ws_conciliacao.cell(row=l_final+1, column=4, value=pagamento.id_orcamento.valor_liquido)
                ws_conciliacao.cell(row=l_final+2, column=1, value='RENDIMENTO DE APLICAÇÃO FINANCEIRA (R$) [B]')
                ws_conciliacao.merge_cells(start_row=l_final+2, start_column=1, end_row=l_final+2, end_column=3)
                ws_conciliacao.cell(row=l_final+2, column=4, value=formulario.cleaned_data['valor_aplicacao'])
                ws_conciliacao.cell(row=l_final+3, column=1, value='TOTAL DA DESPESA (R$) [C]')
                ws_conciliacao.merge_cells(start_row=l_final+3, start_column=1, end_row=l_final+3, end_column=3)
                ws_conciliacao.cell(row=l_final+3, column=4, value='=D{}'.format(l_final-1))
                ws_conciliacao.cell(row=l_final+4, column=1, value='DEVOLUÇÃO / SALDO (R$) [A+B-C]')
                ws_conciliacao.merge_cells(start_row=l_final+4, start_column=1, end_row=l_final+4, end_column=3)
                ws_conciliacao.cell(row=l_final+4, column=4, value='=D{}+D{}-D{}'.format(l_final+1, l_final+2, l_final+3))

                wb.save('media/relatorios.xlsx')

                response = HttpResponse(open('media/relatorios.xlsx', mode='rb', buffering=True), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=relatorios.xlsx'
                return response

        else:
            formulario = FormGerarRelatoriosFinanceiros()
    else:
        formulario = FormGerarRelatoriosFinanceiros()

    return render(request, 'projeto/relatoriosfinanceiros.html', {'projeto':projeto, 'formulario':formulario})

