from projeto.models import Projeto
from .forms import FormUploadArquivo
from django.core.files.storage import FileSystemStorage
from django.db import models
from utils.utilitarios import *
from openpyxl import load_workbook
from .models import *
import xlrd

class Carregaarquivo():

    cabecalho_formulario = '<form method="post" enctype="multipart/form-data"><table class="table">'
    objeto_associado = ""
    rodape_formulario = '</table><button type="submit">Carregar</button></form>'

    def formulario(self):
        if self.objeto_associado:
            form = FormUploadArquivo(initial={'objeto_associado':self.objeto_associado})
        else:
            form = FormUploadArquivo()
        return form

    def model_form_upload(self, request):
        if request.method == 'POST':
            form_upload = FormUploadArquivo(request.POST, request.FILES)
            if form_upload.is_valid():
                #print("id_obj: {} e o Outro: {}".format(request.POST['id_objeto'], form_upload.hidden_fields()))
                #form_upload.objeto_associado = request.POST['id_objeto']
                form_upload.save()
                self.formulario = None
            else:
                self.formulario = form_upload
        return self.formulario

    def caminho_midias(self):
        fs = FileSystemStorage()
        return fs.base_location + "\\"

    def caminho_completo_arquivo(self, nome):
        return self.caminho_midias() + nome

    def arquivo_midia_existe(self, nome):
        fs = FileSystemStorage()
        return True if fs.exists(self.caminho_completo_arquivo(nome)) else False


class ImportaPlanilha():

    planilha_importada = None

    # indicacao de celular na forma do pacote OpenPyXl
    __linha_inicial = 1
    __linha_final = 10
    __coluna_inicial = 'A'
    __coluna_final = 'K'

    # indicacao de celular na forma do pacote xldr
    __idx_linha_inicial = 0
    __idx_linha_final = 9
    __idx_coluna_inicial = 0
    __idx_coluna_final = 9

    # indica para o interpretador se, dentro da regiao indicada, ele deve parar caso haja linha ou coluna vazia
    para_se_linha_vazia = False
    para_se_coluna_vazia = False

    # string do tipo '2-4;8' para indicar que o intervalo da linha 2 a 4 devem ser desconsideradas e, tambem, a linha 8
    # no caso das colunas o formato eh igual, apenas substituidos por letras. Ex.: 'C-F;K'
    __desconsidera_linhas = ""
    __desconsidera_colunas = ""

    # indica a ordem em que funcoes sao aplicadas as colunas. Ex: A=int (converte os dados da coluna em int)
    __aplica_funcoes_colunas = ""

    # indica a(s) funcao(oes) que devem ser aplicadas a blocos ou linhas individuais. Ex: 1-2=etapa('inicial');
    __aplica_funcao_linhas = ""

    def set_linha_inicial(self, linha):
        if isinstance(linha, int) and linha < 0:
            raise ValueError("A linha inicial tem que ter valor maior que zero. Foi informado '{}'.".format(linha))
        self.__linha_inicial = linha
        self.__idx_linha_inicial = linha - 1

    def get_linha_inicial(self):
        return self.__linha_inicial

    def set_linha_final(self, linha):
        if isinstance(linha, int) and linha < 0:
            raise ValueError("A linha final tem que ter valor maior que zero. Foi informado '{}'.".format(linha))
        self.__linha_final = linha
        self.__idx_linha_final = linha - 1

    def get_linha_final(self):
        return self.__linha_final

    def set_coluna_inicial(self, coluna):
        if isinstance(coluna, str):
            coluna = coluna[0].upper()
            if not coluna in a_z_maiusculo():
                raise ValueError("A coluna inicial deve estar entre 'A' e 'Z'. Foi informado '{}'.".format(coluna))
        else:
            raise ValueError("O objeto valor para a coluna inicial não é uma string")
        self.__coluna_inicial = coluna
        self.__idx_coluna_inicial = self.converte_coluna_em_indice(coluna)

    def get_coluna_inicial(self):
        return self.__coluna_inicial

    def set_coluna_final(self, coluna):
        if isinstance(coluna, str):
            coluna = coluna[0].upper()
            if not coluna in a_z_maiusculo():
                raise ValueError("A coluna final deve estar entre 'A' e 'Z'. Foi informado '{}'.".format(coluna))
        else:
            raise ValueError("O objeto valor para a coluna final não é uma string")
        self.__coluna_final = coluna
        self.__idx_coluna_final = self.converte_coluna_em_indice(coluna)

    def get_coluna_final(self):
        return self.__coluna_final

    def set_desconsidera_linhas(self, excecoes):
        self.__desconsidera_linhas = excecoes

    def set_desconsidera_colunas (self, excecoes):
        self.__desconsidera_colunas  = excecoes

    def considerar_todas_linhas(self):
        self.__desconsidera_linhas = ""

    def considerar_todas_colunas(self):
        self.__desconsidera_colunas = ""

    def set_funcoes_linhas(self, funcoes):
        self.__aplica_funcao_linhas = funcoes

    def set_funcoes_colunas(self, funcoes):
        self.__aplica_funcoes_colunas = funcoes

    def ativa_filtro(self, filtro):
        if filtro:
            self.set_linha_inicial(filtro.linha_inicial)
            self.set_linha_final(filtro.linha_final)
            self.set_coluna_inicial(filtro.coluna_inicial)
            self.set_coluna_final(filtro.coluna_final)
            self.set_desconsidera_colunas(filtro.excecao_colunas)
            self.set_desconsidera_linhas(filtro.excecao_linhas)
            self.set_funcoes_colunas(filtro.funcoes_colunas)
            self.set_funcoes_linhas(filtro.funcoes_linhas)


    def desativa_filtro(self):
        self.set_linha_inicial(0)
        self.set_linha_final(10)
        self.set_coluna_inicial('A')
        self.set_coluna_final('K')
        self.set_desconsidera_colunas("")
        self.set_desconsidera_linhas("")
        self.set_funcoes_colunas("")
        self.set_funcoes_linhas("")

    def importa_planilha(self, planilha):

        ca = Carregaarquivo()
        ws = False
        lista_planilha = []
        nome = planilha.arquivo_carga.name
        if ca.arquivo_midia_existe(nome):

            # Verifica se os dados foram inseridos indevidamente
            if (self.__linha_inicial > self.__linha_final) or \
                    (ord(self.__coluna_inicial) > ord(self.__coluna_final)):
                raise ValueError('A linha inicial é maior que a linha final ou a coluna inicial é maior que a coluna final')

            print("LI {} LF{}".format(self.__linha_inicial, self.__linha_final))

            if extensao_arquivo(nome) == 'xlsx':
                wb = load_workbook(ca.caminho_completo_arquivo(nome))
                ws = wb.active
                linhas = range(self.__linha_inicial, self.__linha_final+1)
                colunas = [chr(x) for x in range(ord(self.__coluna_inicial), ord(self.__coluna_final)+1)]
                for linha in linhas:
                    celulas_linhas = []
                    for coluna in colunas:
                        celulas_linhas.append(ws[coluna+str(linha)].value)
                    lista_planilha.append(celulas_linhas)

            if extensao_arquivo(nome) == 'xls':
                # faz a conversao de coluna para poder lidar com a bilioteca xlrd
                wb = xlrd.open_workbook(ca.caminho_completo_arquivo(nome))
                ws = wb.sheet_by_index(0)
                for nr_linha in range(self.__idx_linha_inicial, self.__idx_linha_final+1):
                    linha = ws.row_values(nr_linha)[self.__idx_coluna_inicial:self.__idx_coluna_final + 1]
                    lista_planilha.append(linha)

        return lista_planilha

    def converte_coluna_em_indice(self, str_coluna):
        '''
        Converte colunas indicadas como letras (ex. 'A', 'AD'), para indice de coluna numerico.

        A funcao faz a conversao para manter a compatibilidade de tratamento entre diferentes
        pacotes que tratam da importacao ou expotacao de planilhas Excel
        :param str_coluna: um indice de coluna (ex.: 'A', 'AD')
        :return: indice numerico da coluna iniciando em 0 (zero)
        '''
        indice = 0
        if str_coluna and type(str_coluna) is str:
            base = 26
            multiplicadores = range(1, 27)
            colunas_AZ = [chr(x+64) for x in multiplicadores]
            if len(str_coluna) == 1:
                if str_coluna in colunas_AZ:
                    indice = colunas_AZ.index(str_coluna.upper())
            elif len(str_coluna) == 2:
                str_multiplicacao = str_coluna[0]
                str_soma = str_coluna[1]
                if (str_multiplicacao in colunas_AZ) and (str_soma in colunas_AZ):
                    indice = ((colunas_AZ.index(str_multiplicacao) + 1) * base) + colunas_AZ.index(str_soma)
            else:
                raise IndexError('A função converte_coluna_em_indice() não trata colunas com mais de duas letras. Ex.: AAA')

        return indice


    def regiao_importacao(self, regiao):
        if ":" in regiao:
            celulas_regiao = regiao.split(":")
            celula_inicial = celulas_regiao[0]
            celula_final = celulas_regiao[1]


'''
class ImportaPlanilha():

    linha_inicial = 1
    linha_final = 1
    coluna_inicial = 'A'
    coluna_final = 'A'

    __idx_linha_inicial = 0
    __idx_linha_final = 1
    __idx_coluna_inicial = 0
    __idx_coluna_final = 1

    para_se_linha_vazia = False
    usar_filtro_arquivo = True

    desconsidera_linhas = ""
    desconsidera_colunas = ""

    def importa_planilha(self, arquivo):

        self.coluna_inicial = self.coluna_inicial.upper()
        self.coluna_final = self.coluna_final.upper()

        ca = Carregaarquivo()
        ws = False
        lista_planilha = []
        nome = arquivo.arquivo_carga.name
        if ca.arquivo_midia_existe(nome):

            if arquivo.filtro and self.usar_filtro_arquivo:
                self.linha_inicial = arquivo.filtro.linha_inicial
                self.linha_final = arquivo.filtro.linha_final
                self.coluna_inicial = arquivo.filtro.coluna_inicial
                self.coluna_final = arquivo.filtro.coluna_final

            # Verifica se os dados foram inseridos indevidamente
            if (self.linha_inicial <= 0) or (self.linha_final <= 0):
                raise ValueError('A linha inicial ou a linha final é menor ou igual a zero')
            elif (self.linha_inicial > self.linha_final) or \
                    (ord(self.coluna_inicial.upper()) > ord(self.coluna_final.upper())):
                raise ValueError('A linha inicial é maior que a linha final ou a coluna inicial é maior que a coluna final')

            if extensao_arquivo(nome) == 'xlsx':
                wb = load_workbook(ca.caminho_completo_arquivo(nome))
                ws = wb.active
                linhas = range(self.linha_inicial, self.linha_final+1)
                colunas = [chr(x) for x in range(ord(self.coluna_inicial.upper()), ord(self.coluna_final.upper())+1)]
                for linha in linhas:
                    celulas_linhas = []
                    for coluna in colunas:
                        celulas_linhas.append(ws[coluna+str(linha)].value)
                    lista_planilha.append(celulas_linhas)

            if extensao_arquivo(nome) == 'xls':
                # faz a conversao de coluna para poder lidar com a bilioteca xlrd
                self.__idx_linha_inicial = self.linha_inicial - 1
                self.__idx_linha_final = self.linha_final
                self.__idx_coluna_inicial = self.converte_coluna_em_indice(self.coluna_inicial)
                self.__idx_coluna_final = self.converte_coluna_em_indice(self.coluna_final)
                wb = xlrd.open_workbook(ca.caminho_completo_arquivo(nome))
                ws = wb.sheet_by_index(0)
                for nr_linha in range(self.__idx_linha_inicial, self.__idx_linha_final):
                    linha = ws.row_values(nr_linha)[self.__idx_coluna_inicial:self.__idx_coluna_final + 1]
                    lista_planilha.append(linha)
        return lista_planilha

    def converte_coluna_em_indice(self, str_coluna):
        indice = 0
        if str_coluna and type(str_coluna) is str:
            base = 26
            multiplicadores = range(1, 27)
            colunas_AZ = [chr(x+64) for x in multiplicadores]
            if len(str_coluna) == 1:
                if str_coluna in colunas_AZ:
                    indice = colunas_AZ.index(str_coluna.upper())
            elif len(str_coluna) == 2:
                str_multiplicacao = str_coluna[0]
                str_soma = str_coluna[1]
                if (str_multiplicacao in colunas_AZ) and (str_soma in colunas_AZ):
                    indice = ((colunas_AZ.index(str_multiplicacao) + 1) * base) + colunas_AZ.index(str_soma)
            else:
                raise IndexError('A função converte_coluna_em_indice() não trata colunas com mais de duas letras. Ex.: AAA')

        return indice


    def regiao_importacao(self, regiao):
        if ":" in regiao:
            celulas_regiao = regiao.split(":")
            celula_inicial = celulas_regiao[0]
            celula_final = celulas_regiao[1]
'''
